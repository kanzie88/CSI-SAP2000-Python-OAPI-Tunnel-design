# -*- coding: utf-8 -*-
"""
Created on Fri Aug 11 14:03:35 2023

@author: 10022767
"""

#%% 0) IMPORT REQUIRED MODULES AND DEFINE REQUIRED FILE PATHS

# ================================================================================================================================
# USERDEFINED(to update to your local machine paths)
# indicate section type and internal pile joint label else put None if none

Parent_folder=r"C:\Users\10022767\Surbana Jurong Private Limited(1)\Johanna Enriquez - C-SG-003399_CAG Intra Tunnel\Calculations\CST"
spacing=6 #spacing of piles
# ================================================================================================================================


# 
import pandas as pd
import openpyxl
import xlwings as xw
from xlwings.constants import DeleteShiftDirection
import numpy as np
import os
import win32com.client
import comtypes.client
import pandas as pd
import Getting_aggreagted_loads_sap2000_CST2
import Getting_box_loads5_Ver2
import Extracting_pile_loads2
from itertools import zip_longest
from Extracting_Sap2000_model_info import get_tunnel_sapmodel_info


section_type = 'B2'
Model_path= None

# Access the stored values outside the main loop
print("Stored Section Type:", section_type)
print("SAP2000 Model path(optional):", str(Model_path))



#---------------------------------------------------------------------------------------

# FOR GENERATING BOX LOADS IN df2
# path for tunnel excel calculation as well as the exported LC2 and 51 sheet
loads_cal_path=Parent_folder+"\Section "+section_type+"\Load Calculation and Spring-Section "+section_type+".xlsx"


# for reinf design 
Design_sheet_path_moment=Parent_folder+"\Section "+section_type+"\Axial & Flexural Design-Section "+section_type+".xlsm"
Design_sheet_path_shear=Parent_folder+"\Section "+section_type+"\Shear Design - Section "+section_type+".xlsm"

load_combi_summary_path=Parent_folder+"\Model information\Load combinatios CST.xlsx"
Load_summary_df=pd.read_excel(load_combi_summary_path,sheet_name="Load summary(updated)",header=1,converters={'load_case': str})

# table summary of load_Scenarios to consider
Load_combi_to_include_for_piling=list(Load_summary_df.columns[2:24])
load_combi_to_remove=['UPL-ULS','UPL-SLS'] #only for UPL tension pile design
Load_combi_to_include=Load_combi_to_include_for_piling.copy()
Load_combi_to_include = [item for item in Load_combi_to_include if item not in load_combi_to_remove]

Load_cases_to_include=list(Load_summary_df.load_case)

# Pile_loads_path
Pile_loads_path=Parent_folder+"\Section "+section_type+"\Pile loads.xlsx"


#%% 1) INITIATLIZE MODEL **------------------------- ONLY RUN IF SAP2000 IS NOT OPENED-------------------------**

print("Model Initializing")

SapObject = win32com.client.Dispatch("CSI.SAP2000.API.SapObject")
SapObject.ApplicationStart()
SapModel = SapObject.SapModel  # create SAP2000 model object
SapModel.InitializeNewModel()

if Model_path is None:
    print("Open model from SAP2000 file itself")
    pass
else:
    SapModel.File.OpenFile(Model_path)

SapModel.SetPresentUnits(6)

print("Model Initialized")

#%% 2) EXTRACTING MODEL INFORMATION



model_info_result=get_tunnel_sapmodel_info(SapModel)
df_connectivity=model_info_result[0]
df_frames=model_info_result[1]
df_joint_coords=model_info_result[2]
df_spring_joints=model_info_result[3]
df_joint_coords[["XorR",'Y',"Z"]]=df_joint_coords[["XorR",'Y',"Z"]].round(4)
# identify pile joints
pile_joints=list(df_spring_joints[df_spring_joints.U3>0].Joint)

# filter out pile joints
# identify the outerpiles(min and max X coord)
df_joint_coords_piles=df_joint_coords[df_joint_coords["Joint"].isin(pile_joints)]
df_joint_coords_piles.loc["Outer/Inner"]=np.nan

df_joint_coords_piles.loc[df_joint_coords_piles.XorR==df_joint_coords_piles.XorR.max(),"Outer/Inner"]="Outer"
df_joint_coords_piles.loc[df_joint_coords_piles.XorR==df_joint_coords_piles.XorR.min(),"Outer/Inner"]="Outer"
df_joint_coords_piles.loc[df_joint_coords_piles["Outer/Inner"].isna(),"Outer/Inner"]="Inner"





#%%#3) [BOX LOADS] GETTING BOX LOADS **------------------------- ONLY RUN IF BOX LOADS CHANGED-------------------------**


# Userdefined path for load Load Calculation and Spring-Section A

df2=Getting_box_loads5_Ver2.Creating_box_loads(section_type,loads_cal_path,Joint_coord=df_joint_coords,connectivity=df_connectivity) # requires "Joint Coordinates" and "Connectivity - Frame"


#%% 4) [BOX LOADS] APPLYING TO MODEL BOX LOADS ** ONLY RUN IF BOX LOADS CHANGED**
# =============================================================================

# update model loads using df
SapModel.SetPresentUnits(6)

for i in range(len(df2)):
     
    name=df2.loc[i, "Frame"]
    loadPat=df2.loc[i, "LoadPat"]
    myType=1
    Dir=df2.loc[i, "Dir"]
    Dist1=float(0)
    Dist2=float(1)
    Val1=df2.loc[i, "FOverLA"]
    Val2=df2.loc[i, "FOverLB"]
    CSys="GLOBAL"
    RelDist=True
    Replace=True
    ItemType=int(0)

    
    
    SapModel.FrameObj.SetLoadDistributed(name,loadPat,myType,Dir,Dist1,Dist2,Val1,Val2,CSys,RelDist,Replace,ItemType)
    
    i+=1

print("loads updated")
#%% 5) RUN ANALYSIS

print("Running analysis")
SapModel.Analyze.RunAnalysis()
print("Analysis completed")


#%% 6)[FRAME] EXTRACTING FRAME FORCES
# set to kn,m
SapModel.SetPresentUnits(6)

column_name=['Frame','Station','Elm','ElmSta','OutputCase','StepType','StepNum','P','V2','V3','T','M2','M3']
first_run=True
frame_elements=model_info_result[4]
for i in frame_elements:
    for LoadCombi in Load_combi_to_include:
        SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        SapModel.Results.Setup.SetComboSelectedForOutput(Name=LoadCombi,Selected=True)
        if first_run:
            df3=pd.DataFrame(SapModel.Results.FrameForce(i,ItemTypeElm=0)[2:],index=column_name).transpose()
            first_run=False
        else:
            # df3=df3.append(pd.DataFrame(SapModel.Results.FrameForce(i,ItemTypeElm=0)[2:],index=column_name).transpose())
           
            df3=pd.concat([df3,pd.DataFrame(SapModel.Results.FrameForce(i,ItemTypeElm=0)[2:],index=column_name).transpose()])
df3=df3.loc[df3.OutputCase!='ULS Envelope']    
df3.drop_duplicates(inplace=True)   

print("Frame forces extracted")

# extract the joint loads


#%% 7)[FRAME] AGGREGATING THE FRAME FORCES AND UPDATING DESIGN SPREADSHEET
#
frames_LS_df_moment=pd.read_excel(Design_sheet_path_moment,sheet_name='Summary-input',header=0)[["Frame","Location"]].astype({"Frame":'str'})
frames_LS_df_shear=pd.read_excel(Design_sheet_path_shear,sheet_name='Summary-input',header=0)[["Frame","Location"]].astype({"Frame":'str'})

results=Getting_aggreagted_loads_sap2000_CST2.MOMENT_AND_AXIAL_DESIGN(section_type,df3,frames_LS_df_moment,frames_LS_df_shear)

for dataframe in results:
    dataframe.drop_duplicates(inplace=True)
    
Getting_aggreagted_loads_sap2000_CST2.populating_moment(Design_sheet_path_moment,frames_LS_df_compiled=results[0],frames_LS_df_compiled_SLS=results[1])
Getting_aggreagted_loads_sap2000_CST2.populating_shear(Design_sheet_path_shear,frames_df_compiled=results[2],frames_df_compiled_SLS=results[3])


#%% 8)[PILING] EXTRACTING PILE LOADS AND AGGREGATING
#
column_name= ['Joint','Elm','OutputCase','StepType','StepNum','F1','F2','F3','M1','M2','M3']




# for loadcases
first_run=True
for pile in pile_joints:
    for LoadCases in Load_cases_to_include:
        SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        SapModel.Results.Setup.SetCaseSelectedForOutput(Name=LoadCases,Selected=True)
        if first_run:
            df5=pd.DataFrame(SapModel.Results.JointReact(pile,ItemTypeElm=0)[2:],index=column_name).transpose()
            first_run=False
            
        else:
            # df5=df5.append(pd.DataFrame(SapModel.Results.JointReact(pile,ItemTypeElm=0)[2:],index=column_name).transpose())
            df5=pd.concat([df5,pd.DataFrame(SapModel.Results.JointReact(pile,ItemTypeElm=0)[2:],index=column_name).transpose()])

df5=df5.loc[df5.OutputCase!='ULS Envelope']
df5.drop_duplicates(inplace=True)  
df5.dropna(inplace=True)  

# for load combi
first_run=True
for pile in pile_joints:
    for LoadCombi in Load_combi_to_include_for_piling:
        SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        SapModel.Results.Setup.SetComboSelectedForOutput(Name=LoadCombi,Selected=True)
        if first_run:
            df4=pd.DataFrame(SapModel.Results.JointReact(pile,ItemTypeElm=0)[2:],index=column_name).transpose()
            first_run=False
            
        else:
            # df4=df4.append(pd.DataFrame(SapModel.Results.JointReact(pile,ItemTypeElm=0)[2:],index=column_name).transpose())
            df4=pd.concat([df4,pd.DataFrame(SapModel.Results.JointReact(pile,ItemTypeElm=0)[2:],index=column_name).transpose()])

df4=df4.loc[df4.OutputCase!='ULS Envelope']    
df4.drop_duplicates(inplace=True)
df4.dropna(inplace=True) 




final=Extracting_pile_loads2.extract_pile_loads(df_loadcases=df5,df_loadcombi=df4,Load_summary_df=Load_summary_df,df_joint_coords_piles=df_joint_coords_piles)
# add column for pile load for pile spaced at xx

final[0][f'unfactored load per pile spaced at {spacing}m']=final[0]['unfactored-per m run']*spacing

if isinstance(final[1], pd.DataFrame):
            final[1][f'unfactored load per pile spaced at {spacing}m']=final[1]['unfactored-per m run']*spacing
        
            

#%% 9) [PILING] WRITING PILE LOADS
from openpyxl import load_workbook
# Load the source and destination Excel workbooks
source_wb = openpyxl.load_workbook(Pile_loads_path)


try:
    # Get the source and destination sheets
    source_sheet_name = 'Sheet1' 
    dest_sheet_name = 'Sheet1(Previous)'  
    source_sheet = source_wb[source_sheet_name]
    dest_sheet = source_wb[dest_sheet_name]
    
except:
    
    wb2 = load_workbook(Pile_loads_path)
    wb2.create_sheet('Sheet1(Previous)')
    wb2.save(Pile_loads_path)
    
for row in dest_sheet.iter_rows():
    for cell in row:
        cell.value = None

# Iterate through the source sheet's cells and copy them to the destination sheet
for row in source_sheet.iter_rows():
    for cell in row:
        dest_sheet[cell.coordinate].value = cell.value
        print(cell.value)



# updating pile loads
wb = xw.Book(Pile_loads_path)
ws = wb.sheets['Sheet1']
print("Clearing previous loads")
ws["A2:G1000"].clear()
ws["A2"].options(pd.DataFrame,headers=True, index=False, expand='table').value = final[0]
if isinstance(final[1], pd.DataFrame):
            ws["A10"].value="Centre pile"
            final[1][f'unfactored load per pile spaced at {spacing}m']=final[1]['unfactored-per m run']*spacing
            ws["A11"].options(pd.DataFrame,headers=True, index=False, expand='table').value = final[1]
            
print("Updated loads")



#%% 10)[PILING] GET DISPLACEMENT
columns=["Joint","LoadCombi","U1","U2","U3"]
df6 = pd.DataFrame(columns=columns)
Load_combi_to_include_serv=[item for item in Load_combi_to_include if item.startswith("2")]
for pile in list(pile_joints.Joints):
    for LoadCombi in Load_combi_to_include_serv:
        SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        SapModel.Results.Setup.SetComboSelectedForOutput(Name=LoadCombi,Selected=True)
        # df6=df6.append(df6_temp)
        # df6_temp=df6_temp[0:0]
        new_row = {
                   "LoadCombi": LoadCombi, 
                   'Joint': pile, 
                   'UX': SapModel.Results.JointDispl(pile,ItemTypeElm=0)[7][0],
                   'UY': SapModel.Results.JointDispl(pile,ItemTypeElm=0)[8][0],
                   'UZ': SapModel.Results.JointDispl(pile,ItemTypeElm=0)[9][0]
                   }
        df6 = df6.append(new_row, ignore_index=True)

#%% 10) EXIT PROGRAMME
SapObject.ApplicationExit(False)
     